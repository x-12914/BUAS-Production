"""
Device Data Excel Export Service
Generates Excel files with device-specific data across multiple tabs.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime, date
import io
import pytz
from flask import current_app
from ..models import db, DeviceLocation, RecordingEvent, Upload, SmsMessage, CallLog, NIGERIAN_TZ
from ..device_utils import resolve_to_device_id
from ..utils.audio_file_resolver import resolve_audio_file

class DeviceExcelExporter:
    """Handles Excel export for device-specific data"""
    
    def __init__(self):
        self.workbook = None
        self.device_id = None
        self.device_name = None
        
    def export_device_data(self, device_id, start_date=None, end_date=None):
        """
        Export all device data to Excel with multiple tabs
        
        Args:
            device_id (str): Device identifier
            start_date (datetime, optional): Filter start date
            end_date (datetime, optional): Filter end date
            
        Returns:
            io.BytesIO: Excel file as bytes
        """
        # Resolve device ID
        self.device_id = resolve_to_device_id(device_id)
        
        # Get device display name
        from ..models import DeviceInfo
        device_info = DeviceInfo.query.filter_by(device_id=self.device_id).first()
        self.device_name = device_info.display_name if device_info and device_info.display_name else self.device_id
        
        # Create workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Export each data type
        locations_count = self._export_locations(start_date, end_date)
        recordings_count = self._export_recordings(start_date, end_date) 
        contacts_count = self._export_contacts(start_date, end_date)
        sms_count = self._export_sms(start_date, end_date)
        call_logs_count = self._export_call_logs(start_date, end_date)
        
        # Add summary sheet as first tab
        self._add_summary_sheet(locations_count, recordings_count, contacts_count, sms_count, call_logs_count, start_date, end_date)
        
        # Move summary to front
        summary_sheet = self.workbook['Summary']
        self.workbook.move_sheet(summary_sheet, -(len(self.workbook.sheetnames) - 1))
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        self.workbook.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _export_locations(self, start_date=None, end_date=None):
        """Export device locations to Excel tab"""
        
        # Query locations
        query = DeviceLocation.query.filter_by(device_id=self.device_id)
        
        if start_date:
            query = query.filter(DeviceLocation.timestamp >= start_date)
        if end_date:
            query = query.filter(DeviceLocation.timestamp <= end_date)
            
        locations = query.order_by(DeviceLocation.timestamp.desc()).all()
        
        # Create worksheet
        ws = self.workbook.create_sheet("Locations")
        
        # Headers
        headers = [
            'Date', 'Time', 'Latitude', 'Longitude', 'Timestamp'
        ]
        
        # Style headers
        self._add_headers(ws, headers)
        
        # Add data
        for row_num, location in enumerate(locations, start=2):
            ws[f'A{row_num}'] = location.date.strftime('%Y-%m-%d') if location.date else 'N/A'
            ws[f'B{row_num}'] = location.time.strftime('%H:%M:%S') if location.time else 'N/A'
            ws[f'C{row_num}'] = location.latitude
            ws[f'D{row_num}'] = location.longitude
            ws[f'E{row_num}'] = location.timestamp.strftime('%Y-%m-%d %H:%M:%S') if location.timestamp else 'N/A'
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        return len(locations)
    
    def _export_recordings(self, start_date=None, end_date=None):
        """Export recording events to Excel tab"""
        
        # Query recordings
        query = RecordingEvent.query.filter_by(device_id=self.device_id)
        
        if start_date:
            query = query.filter(RecordingEvent.start_date >= start_date.date() if hasattr(start_date, 'date') else start_date)
        if end_date:
            query = query.filter(RecordingEvent.start_date <= end_date.date() if hasattr(end_date, 'date') else end_date)
            
        recordings = query.order_by(RecordingEvent.start_date.desc()).all()
        
        # Create worksheet
        ws = self.workbook.create_sheet("Recordings")
        
        # Headers
        headers = [
            'Start Date', 'Start Time', 'Stop Date', 'Stop Time', 
            'Duration (mins)', 'Status', 'Audio File', 'Audio Link',
            'Start Location', 'Stop Location'
        ]
        
        # Style headers
        self._add_headers(ws, headers)
        
        # Add data
        for row_num, recording in enumerate(recordings, start=2):
            # Convert UTC times to Nigerian time for display
            start_nigerian = None
            stop_nigerian = None
            
            if recording.start_date and recording.start_time:
                start_utc = datetime.combine(recording.start_date, recording.start_time).replace(tzinfo=pytz.utc)
                start_nigerian = start_utc.astimezone(NIGERIAN_TZ)
            
            if recording.stop_date and recording.stop_time:
                stop_utc = datetime.combine(recording.stop_date, recording.stop_time).replace(tzinfo=pytz.utc)
                stop_nigerian = stop_utc.astimezone(NIGERIAN_TZ)
            
            # Calculate duration using Nigerian times
            duration = None
            if start_nigerian and stop_nigerian:
                try:
                    duration = (stop_nigerian - start_nigerian).total_seconds() / 60  # minutes
                except:
                    duration = None
            
            # Use Nigerian times for display
            ws[f'A{row_num}'] = start_nigerian.strftime('%Y-%m-%d') if start_nigerian else 'N/A'
            ws[f'B{row_num}'] = start_nigerian.strftime('%H:%M:%S') if start_nigerian else 'N/A'
            ws[f'C{row_num}'] = stop_nigerian.strftime('%Y-%m-%d') if stop_nigerian else 'N/A'
            ws[f'D{row_num}'] = stop_nigerian.strftime('%H:%M:%S') if stop_nigerian else 'N/A'
            ws[f'E{row_num}'] = f"{duration:.1f}" if duration else 'N/A'
            ws[f'F{row_num}'] = 'Completed' if recording.stop_date else 'In Progress'
            
            # Use the new audio file resolver to find actual files
            start_date_str = start_nigerian.strftime('%Y-%m-%d') if start_nigerian else None
            start_time_str = start_nigerian.strftime('%H:%M:%S') if start_nigerian else None
            
            audio_filename, audio_url = resolve_audio_file(
                device_id=self.device_id,
                audio_file_id=recording.audio_file_id,
                start_date=start_date_str,
                start_time=start_time_str,
                base_url="http://105.114.25.157:5000"
            )
            
            # Set the audio file name
            ws[f'G{row_num}'] = audio_filename or 'N/A'
            
            # Audio link (hyperlinked)
            if audio_filename and audio_url != "Audio Not Available":
                cell = ws[f'H{row_num}']
                cell.value = audio_url
                cell.hyperlink = Hyperlink(ref=f"H{row_num}", target=audio_url)
                cell.font = Font(color="0000FF", underline="single")
            else:
                ws[f'H{row_num}'] = 'No audio file found'
            
            # Location data (shifted columns)
            ws[f'I{row_num}'] = f"{recording.start_latitude}, {recording.start_longitude}" if recording.start_latitude and recording.start_longitude else 'N/A'
            ws[f'J{row_num}'] = f"{recording.stop_latitude}, {recording.stop_longitude}" if recording.stop_latitude and recording.stop_longitude else 'N/A'
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        return len(recordings)
    
    def _export_contacts(self, start_date=None, end_date=None):
        """Export phone contacts to Excel tab"""
        
        # Get contacts from device_info table (stored as JSON)
        from ..models import DeviceInfo
        device_info = DeviceInfo.query.filter_by(device_id=self.device_id).first()
        
        contacts = []
        if device_info and device_info.contacts:
            try:
                import json
                contacts_data = json.loads(device_info.contacts)
                if isinstance(contacts_data, list):
                    for i, contact in enumerate(contacts_data):
                        if isinstance(contact, dict):
                            contacts.append({
                                'id': i + 1,
                                'name': contact.get('name', 'N/A'),
                                'phone': contact.get('phone', 'N/A'),
                                'created_at': device_info.updated_at or device_info.created_at
                            })
            except (json.JSONDecodeError, Exception) as e:
                current_app.logger.warning(f"Error parsing contacts for device {self.device_id}: {e}")
        
        # Filter by date if provided
        if start_date or end_date:
            filtered_contacts = []
            for contact in contacts:
                contact_date = contact['created_at']
                if contact_date:
                    if start_date and contact_date.date() < start_date.date():
                        continue
                    if end_date and contact_date.date() > end_date.date():
                        continue
                filtered_contacts.append(contact)
            contacts = filtered_contacts
        
        # Create worksheet
        ws = self.workbook.create_sheet("Contacts")
        
        # Headers
        headers = [
            'Contact ID', 'Name', 'Phone Number', 'Sync Date'
        ]
        
        # Style headers
        self._add_headers(ws, headers)
        
        # Add data
        for row_num, contact in enumerate(contacts, start=2):
            ws[f'A{row_num}'] = contact['id']
            ws[f'B{row_num}'] = contact['name']
            ws[f'C{row_num}'] = contact['phone']
            ws[f'D{row_num}'] = contact['created_at'].strftime('%Y-%m-%d %H:%M:%S') if contact['created_at'] else 'N/A'
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        return len(contacts)
    
    def _export_sms(self, start_date=None, end_date=None):
        """Export SMS messages to Excel tab"""
        
        # Query SMS messages (received messages only)
        query = SmsMessage.query.filter_by(device_id=self.device_id, direction='inbox')
        
        if start_date:
            query = query.filter(SmsMessage.date >= start_date)
        if end_date:
            query = query.filter(SmsMessage.date <= end_date)
            
        sms_messages = query.order_by(SmsMessage.date.desc()).all()
        
        # Create worksheet
        ws = self.workbook.create_sheet("SMS Messages")
        
        # Headers
        headers = [
            'Date', 'Time', 'From', 'Message', 'Read Status', 'SMS ID'
        ]
        
        # Style headers
        self._add_headers(ws, headers)
        
        # Add data
        for row_num, sms in enumerate(sms_messages, start=2):
            ws[f'A{row_num}'] = sms.date.strftime('%Y-%m-%d') if sms.date else 'N/A'
            ws[f'B{row_num}'] = sms.date.strftime('%H:%M:%S') if sms.date else 'N/A'
            ws[f'C{row_num}'] = sms.address or 'N/A'
            ws[f'D{row_num}'] = sms.body or 'N/A'
            ws[f'E{row_num}'] = 'Read' if sms.read else 'Unread'
            ws[f'F{row_num}'] = sms.sms_id or 'N/A'
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        return len(sms_messages)
    
    def _export_call_logs(self, start_date=None, end_date=None):
        """Export call logs to Excel tab"""
        
        # Query call logs
        query = CallLog.query.filter_by(device_id=self.device_id)
        
        if start_date:
            query = query.filter(CallLog.call_date >= start_date)
        if end_date:
            query = query.filter(CallLog.call_date <= end_date)
            
        call_logs = query.order_by(CallLog.call_date.desc()).all()
        
        # Create worksheet
        ws = self.workbook.create_sheet("Call Logs")
        
        # Headers
        headers = [
            'Date', 'Time', 'Phone Number', 'Contact Name', 'Call Type', 'Duration (seconds)', 'Duration (formatted)'
        ]
        
        # Style headers
        self._add_headers(ws, headers)
        
        # Add data
        for row_num, call in enumerate(call_logs, start=2):
            ws[f'A{row_num}'] = call.call_date.strftime('%Y-%m-%d') if call.call_date else 'N/A'
            ws[f'B{row_num}'] = call.call_date.strftime('%H:%M:%S') if call.call_date else 'N/A'
            ws[f'C{row_num}'] = call.phone_number or 'N/A'
            ws[f'D{row_num}'] = call.contact_name or 'N/A'
            ws[f'E{row_num}'] = call.get_call_type_display()
            ws[f'F{row_num}'] = call.duration or 0
            ws[f'G{row_num}'] = call.format_duration()
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        return len(call_logs)
    
    def _add_summary_sheet(self, locations_count, recordings_count, contacts_count, sms_count, call_logs_count, start_date, end_date):
        """Add summary information sheet"""
        
        ws = self.workbook.create_sheet("Summary")
        
        # Title
        ws['A1'] = f"Device Data Export Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left')
        
        # Device info
        ws['A3'] = "Device Information"
        ws['A3'].font = Font(size=12, bold=True)
        ws['A4'] = f"Device ID: {self.device_id}"
        ws['A5'] = f"Device Name: {self.device_name}"
        ws['A6'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Date range
        ws['A8'] = "Date Range"
        ws['A8'].font = Font(size=12, bold=True)
        ws['A9'] = f"Start Date: {start_date.strftime('%Y-%m-%d') if start_date else 'All Time'}"
        ws['A10'] = f"End Date: {end_date.strftime('%Y-%m-%d') if end_date else 'All Time'}"
        
        # Data summary
        ws['A12'] = "Data Summary"
        ws['A12'].font = Font(size=12, bold=True)
        ws['A13'] = f"Locations: {locations_count:,} records"
        ws['A14'] = f"Recordings: {recordings_count:,} records"
        ws['A15'] = f"Contacts: {contacts_count:,} records"
        ws['A16'] = f"SMS Messages: {sms_count:,} records"
        ws['A17'] = f"Call Logs: {call_logs_count:,} records"
        ws['A18'] = f"Total Records: {(locations_count + recordings_count + contacts_count + sms_count + call_logs_count):,}"
        
        # Instructions
        ws['A20'] = "Instructions"
        ws['A20'].font = Font(size=12, bold=True)
        ws['A21'] = "• Click on the tabs below to view different data types"
        ws['A22'] = "• Locations: GPS coordinates and movement data"
        ws['A23'] = "• Recordings: Audio recording sessions and metadata (click Audio Link to download)"
        ws['A24'] = "• Contacts: Phone contacts from the device"
        ws['A25'] = "• SMS Messages: Text messages received by the device"
        ws['A26'] = "• Call Logs: Phone call history and metadata"
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _add_headers(self, worksheet, headers):
        """Add styled headers to worksheet"""
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set minimum and maximum widths
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

# Convenience function for easy import
def export_device_to_excel(device_id, start_date=None, end_date=None):
    """
    Export device data to Excel file
    
    Args:
        device_id (str): Device identifier
        start_date (datetime, optional): Filter start date
        end_date (datetime, optional): Filter end date
        
    Returns:
        io.BytesIO: Excel file as bytes
    """
    exporter = DeviceExcelExporter()
    return exporter.export_device_data(device_id, start_date, end_date)